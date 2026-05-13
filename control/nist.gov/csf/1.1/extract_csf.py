#!/usr/bin/env python3
"""Extract NIST CSF v1.1 Core from the official XLSX into flat CSV/JSON.

Input: csf-1.1-core.xlsx (sheet "Core", 4 columns).

Source structure: Function | Category | Subcategory | Informative References
- Function row: only Function column populated
- Category row: only Category column populated (Function inherited)
- Subcategory row: Subcategory column populated (Function + Category inherited)
- Continuation rows: only Informative References populated (other 3 inherited)

Each Subcategory may have many Informative References. We aggregate them
into a single field per Subcategory.

Output is one row per Subcategory.
"""
import csv
import json
import re

import openpyxl


def split_function(text):
    """'IDENTIFY (ID): description' is not the format here — v1.1 uses just the name.
    Function cells contain only the function name (sometimes with trailing whitespace).
    Convert 'IDENTIFY ' to ('IDENTIFY', 'ID') by lookup, then derive description elsewhere if needed.
    """
    text = text.strip()
    # Subcategory IDs like ID.AM-1 give us the function ID
    return text, ""


def split_category(text):
    """'Asset Management (ID.AM): description' -> ('Asset Management', 'ID.AM', 'description')"""
    m = re.match(r"^(.+?)\s*\(([A-Z]+\.[A-Z]+)\):\s*(.*)", text, re.DOTALL)
    return m.groups() if m else (text.strip(), "", "")


def split_subcategory(text):
    """'ID.AM-1: description' -> ('ID.AM-1', 'description')"""
    m = re.match(r"^([A-Z]+\.[A-Z]+-\d+):\s*(.*)", text, re.DOTALL)
    return m.groups() if m else ("", text)


def clean_ref(s):
    """Clean an Informative Reference cell."""
    if not s:
        return ""
    # The XLSX uses '· ' (middle dot + non-breaking spaces) as a bullet prefix
    s = re.sub(r"^[·•]\s*", "", s).strip()
    # Collapse non-breaking spaces
    s = s.replace(" ", " ")
    return re.sub(r"\s+", " ", s).strip()


def extract():
    wb = openpyxl.load_workbook("csf-1.1-core.xlsx", data_only=True)
    ws = wb["Core"]

    rows = []
    fn_name = fn_id = ""
    cat_name = cat_id = cat_desc = ""
    sub_id = sub_text = ""
    refs = []

    def flush():
        if sub_id:
            rows.append({
                "Function ID": fn_id,
                "Function Name": fn_name,
                "Category ID": cat_id,
                "Category Name": cat_name,
                "Category Description": cat_desc,
                "Subcategory ID": sub_id,
                "Subcategory": sub_text.strip(),
                "Informative References": "\n".join(refs),
            })

    # v1.1 function name -> ID lookup (the 5 CSF v1.1 functions)
    FN_LOOKUP = {
        "IDENTIFY": "ID",
        "PROTECT": "PR",
        "DETECT": "DE",
        "RESPOND": "RS",
        "RECOVER": "RC",
    }

    # Data starts on row 2 (row 1 is headers)
    for r in range(2, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value

        if c1:
            # New function — flush prior subcategory, reset context.
            # Function cells in v1.1 XLSX have inconsistent format:
            # row 2 has 'IDENTIFY ' (bare name), rows 108/281/344/389 have
            # 'PROTECT (PR)', 'DETECT (DE)', 'RESPOND (RS)', 'RECOVER (RC)'.
            flush()
            raw = c1.strip()
            m = re.match(r"^([A-Z]+)\s*(?:\(([A-Z]+)\))?", raw)
            if m:
                fn_name = m.group(1)
                fn_id = m.group(2) or FN_LOOKUP.get(fn_name.upper(), "")
            else:
                fn_name = raw
                fn_id = FN_LOOKUP.get(fn_name.upper(), "")
            cat_name = cat_id = cat_desc = ""
            sub_id = sub_text = ""
            refs = []

        if c2:
            # New category — flush prior subcategory
            flush()
            cat_name_part, cat_id_part, cat_desc_part = split_category(c2)
            cat_name = cat_name_part
            cat_id = cat_id_part
            cat_desc = cat_desc_part.strip()
            sub_id = sub_text = ""
            refs = []

        if c3:
            # New subcategory — flush prior, start fresh
            flush()
            sub_id, sub_text = split_subcategory(c3)
            refs = []

        if c4:
            ref = clean_ref(c4)
            if ref:
                refs.append(ref)

    # Last subcategory
    flush()
    return rows


def main():
    rows = extract()
    print(f"Extracted {len(rows)} subcategories")

    fields = list(rows[0].keys())
    with open("csf-1.1.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)

    with open("csf-1.1.json", "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)

    from collections import Counter
    fns = Counter(r["Function ID"] for r in rows)
    cats = sorted({r["Category ID"] for r in rows})
    print(f"Functions ({len(fns)}): {dict(fns)}")
    print(f"Categories: {len(cats)} ({cats[:5]}...)")
    print("\nFirst 3 subcategories:")
    for r in rows[:3]:
        print(f"  {r['Subcategory ID']}: {r['Subcategory'][:60]}")
        print(f"    refs: {r['Informative References'].count(chr(10)) + 1 if r['Informative References'] else 0}")


if __name__ == "__main__":
    main()
