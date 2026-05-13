#!/usr/bin/env python3
"""Extract NIST CSF 2.0 Core from the NIST Reference Tool XLSX into flat CSV/JSON.

Input: csf-2.0-core.xlsx (sheet "CSF 2.0", headers on row 2).

Each source row is one of three types, distinguished by which column is populated:
- Function row    → col 1 has "FUNCTION_NAME (FN_ID): description"
- Category row    → col 2 has "Category Name (FN_ID.CAT): description"
- Subcategory row → col 3 has "FN_ID.CAT-NN: subcategory text"

Output is flattened so each Subcategory carries its parent Function/Category.
"""
import csv
import json
import re

import openpyxl


def split_function(text):
    """'GOVERN (GV): description' → ('GOVERN', 'GV', 'description')"""
    m = re.match(r"^([A-Z]+)\s*\(([A-Z]+)\):\s*(.*)", text, re.DOTALL)
    return m.groups() if m else (text, "", "")


def split_category(text):
    """'Organizational Context (GV.OC): description' → ('Organizational Context', 'GV.OC', 'description')"""
    m = re.match(r"^(.+?)\s*\(([A-Z]+\.[A-Z]+)\):\s*(.*)", text, re.DOTALL)
    return m.groups() if m else (text, "", "")


def split_subcategory(text):
    """'GV.OC-01: subcategory text' → ('GV.OC-01', 'subcategory text')"""
    m = re.match(r"^([A-Z]+\.[A-Z]+-\d+):\s*(.*)", text, re.DOTALL)
    return m.groups() if m else ("", text)


def extract():
    wb = openpyxl.load_workbook("csf-2.0-core.xlsx", data_only=True)
    ws = wb["CSF 2.0"]

    rows = []
    fn_name = fn_id = fn_desc = ""
    cat_name = cat_id = cat_desc = ""

    # Headers on row 2; data starts row 3.
    for r in range(3, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        c5 = ws.cell(r, 5).value

        if c1:
            fn_name, fn_id, fn_desc = split_function(c1.strip())
            cat_name = cat_id = cat_desc = ""
        elif c2:
            cat_name, cat_id, cat_desc = split_category(c2.strip())
        elif c3:
            sub_id, sub_text = split_subcategory(c3.strip())
            sub_text = sub_text.strip()
            cat_desc_clean = cat_desc.strip()
            # NIST retains withdrawn CSF 1.1 entries with "[Withdrawn: ...]" markers
            # that point to where the content moved in CSF 2.0.
            status = "Withdrawn" if sub_text.startswith("[Withdrawn:") or cat_desc_clean.startswith("[Withdrawn:") else "Active"
            rows.append({
                "Status": status,
                "Function ID": fn_id,
                "Function Name": fn_name,
                "Function Description": fn_desc.strip(),
                "Category ID": cat_id,
                "Category Name": cat_name,
                "Category Description": cat_desc_clean,
                "Subcategory ID": sub_id,
                "Subcategory": sub_text,
                "Implementation Examples": (c4 or "").strip(),
                "Informative References": (c5 or "").strip(),
            })

    return rows


def main():
    rows = extract()
    print(f"Extracted {len(rows)} subcategories")

    # CSV
    fields = list(rows[0].keys())
    with open("csf-2.0.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)

    # JSON
    with open("csf-2.0.json", "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)

    # Summary
    active = [r for r in rows if r["Status"] == "Active"]
    withdrawn = [r for r in rows if r["Status"] == "Withdrawn"]
    fns = sorted({r["Function ID"] for r in active})
    active_cats = sorted({r["Category ID"] for r in active})
    print(f"Active subcategories: {len(active)}")
    print(f"Withdrawn (CSF 1.1) subcategories: {len(withdrawn)}")
    print(f"CSF 2.0 Functions ({len(fns)}): {fns}")
    print(f"CSF 2.0 Categories: {len(active_cats)}")


if __name__ == "__main__":
    main()
