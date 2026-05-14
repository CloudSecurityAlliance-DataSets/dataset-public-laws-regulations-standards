#!/usr/bin/env python3
"""Extract NIST SP 800-53A Rev 5 assessment procedures from the NIST-published XLSX.

Input: 800-53a-assessment-procedures.xlsx
       (from https://csrc.nist.gov/files/pubs/sp/800/53/a/r5/final/docs/
        sp800-53Ar5-assessment-procedures.xlsx)

Each row is one assessment-objective row for a control. Columns:
  - family             Family name (e.g., "ACCESS CONTROL")
  - identifier         Control ID like AC-01 or AC-02(01)
  - sort-as            Sortable form (e.g., AC-01-00-00, AC-02-01-00)
  - control-name       Title
  - assessment-objective   "Determine if:" header for the row
  - EXAMINE            Things to examine for this objective
  - INTERVIEW          People/orgs to interview
  - TEST               Tests to perform

Controls have multiple assessment objectives (rows). Output records
each row, plus derives family code, base_control, enhancement number.

Output:
  - 800-53a-assessment-procedures.csv
  - 800-53a-assessment-procedures.json
"""
import csv
import json
import re

import openpyxl


def parse_id(identifier: str):
    """AC-01 -> family=AC, base=AC-01, enhancement=None
       AC-02(01) -> family=AC, base=AC-02, enhancement=1"""
    s = (identifier or "").strip()
    m = re.match(r"^([A-Z]{2})-(\d+)(?:\((\d+)\))?", s)
    if not m:
        return "", s, None
    fam = m.group(1)
    num = m.group(2)
    enh = int(m.group(3)) if m.group(3) else None
    return fam, f"{fam}-{num}", enh


def clean(s):
    if s is None:
        return ""
    return re.sub(r"\s+\n", "\n", str(s)).strip()


def main():
    wb = openpyxl.load_workbook("800-53a-assessment-procedures.xlsx", data_only=True)
    # Find the assessment-procedures sheet (skip README)
    sheets = [s for s in wb.sheetnames if s.lower() != "readme"]
    ws = wb[sheets[0]]

    # Header row
    headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    idx = {h.lower(): i for i, h in enumerate(headers)}

    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if not any(vals):
            continue
        identifier = vals[idx["identifier"]]
        family, base, enh = parse_id(identifier)
        rows.append({
            "identifier": identifier,
            "family": family,
            "family_long_name": vals[idx["family"]],
            "base_control": base,
            "enhancement": enh if enh is not None else "",
            "is_enhancement": enh is not None,
            "sort_as": vals[idx["sort-as"]],
            "control_name": vals[idx["control-name"]],
            "assessment_objective": vals[idx["assessment-objective"]],
            "examine": vals[idx["examine"]],
            "interview": vals[idx["interview"]],
            "test": vals[idx["test"]],
        })

    # Sort by family, base number, enhancement, sort-as (for sub-objectives)
    def sort_key(r):
        fam = r["family"]
        m = re.match(r"^[A-Z]{2}-(\d+)", r["base_control"])
        base_num = int(m.group(1)) if m else 999
        enh = r["enhancement"] if r["enhancement"] != "" else 0
        return (fam, base_num, enh, r["sort_as"])

    rows.sort(key=sort_key)

    fields = ["identifier", "family", "family_long_name", "base_control",
              "enhancement", "is_enhancement", "sort_as", "control_name",
              "assessment_objective", "examine", "interview", "test"]

    with open("800-53a-assessment-procedures.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, quoting=csv.QUOTE_ALL)
        w.writeheader()
        w.writerows(rows)

    with open("800-53a-assessment-procedures.json", "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)

    from collections import Counter
    by_family = Counter(r["family"] for r in rows)
    distinct_controls = len({r["identifier"] for r in rows})
    print(f"Total assessment-objective rows: {len(rows)}")
    print(f"Distinct controls/enhancements:  {distinct_controls}")
    print(f"Families ({len(by_family)}):")
    for fam, n in sorted(by_family.items()):
        print(f"  {fam}: {n} rows")


if __name__ == "__main__":
    main()
