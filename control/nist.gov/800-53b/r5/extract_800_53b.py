#!/usr/bin/env python3
"""Extract NIST SP 800-53B control baselines from the NIST-published XLSX.

Input: 800-53b-baselines.xlsx
       (from https://csrc.nist.gov/files/pubs/sp/800/53/b/upd1/final/docs/
        sp800-53B-control-baselines.xlsx)

One row per control or control enhancement, columns indicating which
baseline(s) it belongs to:
  - SORT-AS
  - Control Identifier         (AC-1, AC-2(1), etc.)
  - Control (or Control Enhancement) Name
  - Withdrawn                   ("Yes" if withdrawn)
  - Privacy Baseline            ("x" if in baseline)
  - Security Control Baseline - Low
  - Security Control Baseline - Moderate
  - Security Control Baseline - High

Output:
  - 800-53b-baselines.csv
  - 800-53b-baselines.json
"""
import csv
import json
import re

import openpyxl


def parse_id(identifier: str):
    s = (identifier or "").strip()
    m = re.match(r"^([A-Z]{2})-(\d+)(?:\((\d+)\))?", s)
    if not m:
        return "", s, None
    fam, num = m.group(1), m.group(2)
    enh = int(m.group(3)) if m.group(3) else None
    return fam, f"{fam}-{num}", enh


def clean(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def yn(s):
    """Convert baseline membership ('x', 'x ', '', None) to bool."""
    return clean(s).lower() == "x"


def main():
    wb = openpyxl.load_workbook("800-53b-baselines.xlsx", data_only=True)
    sheets = [s for s in wb.sheetnames if s.lower() != "readme"]
    ws = wb[sheets[0]]
    headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    # Map columns by header
    def col(prefix):
        for i, h in enumerate(headers):
            if h.lower().startswith(prefix.lower()):
                return i
        return None

    c_sort = col("SORT-AS")
    c_id = col("Control Identifier")
    c_name = col("Control (or Control")
    c_withdrawn = col("Withdrawn")
    c_priv = col("Privacy Baseline")
    c_low = col("Security Control Baseline - Lo")
    c_mod = col("Security Control Baseline - Mo")
    c_high = col("Security Control Baseline - Hi")

    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(vals):
            continue
        identifier = clean(vals[c_id])
        if not identifier:
            continue
        fam, base, enh = parse_id(identifier)
        rows.append({
            "identifier": identifier,
            "family": fam,
            "base_control": base,
            "enhancement": enh if enh is not None else "",
            "is_enhancement": enh is not None,
            "sort_as": clean(vals[c_sort]),
            "name": clean(vals[c_name]),
            "withdrawn": clean(vals[c_withdrawn]).lower() == "yes",
            "in_privacy_baseline": yn(vals[c_priv]),
            "in_low_baseline": yn(vals[c_low]),
            "in_moderate_baseline": yn(vals[c_mod]),
            "in_high_baseline": yn(vals[c_high]),
        })

    def sort_key(r):
        fam = r["family"]
        m = re.match(r"^[A-Z]{2}-(\d+)", r["base_control"])
        base_num = int(m.group(1)) if m else 999
        enh = r["enhancement"] if r["enhancement"] != "" else 0
        return (fam, base_num, enh)

    rows.sort(key=sort_key)

    fields = ["identifier", "family", "base_control", "enhancement",
              "is_enhancement", "sort_as", "name", "withdrawn",
              "in_privacy_baseline", "in_low_baseline",
              "in_moderate_baseline", "in_high_baseline"]

    with open("800-53b-baselines.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, quoting=csv.QUOTE_ALL)
        w.writeheader()
        w.writerows(rows)

    with open("800-53b-baselines.json", "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)

    n_low = sum(r["in_low_baseline"] for r in rows)
    n_mod = sum(r["in_moderate_baseline"] for r in rows)
    n_high = sum(r["in_high_baseline"] for r in rows)
    n_priv = sum(r["in_privacy_baseline"] for r in rows)
    n_withdrawn = sum(r["withdrawn"] for r in rows)
    print(f"Total entries: {len(rows)}")
    print(f"  Low baseline:      {n_low}")
    print(f"  Moderate baseline: {n_mod}")
    print(f"  High baseline:     {n_high}")
    print(f"  Privacy baseline:  {n_priv}")
    print(f"  Withdrawn:         {n_withdrawn}")


if __name__ == "__main__":
    main()
