#!/usr/bin/env python3
"""Extract NIST SP 800-53 Rev 5 controls from the NIST-published XLSX.

Input: 800-53-r5-controls.xlsx
       (downloaded from
        https://csrc.nist.gov/files/pubs/sp/800/53/r5/upd1/final/docs/sp800-53r5-controls.xlsx)

The XLSX is the authoritative structured form maintained by NIST alongside
the SP 800-53r5 publication. Each row is either a base control (e.g.,
AC-1) or a control enhancement (e.g., AC-2(1)). Columns:
  - Control Identifier
  - Control (or Control Enhancement) Name
  - Control (or Control Enhancement) [text]
  - Discussion
  - Related Controls

This script flattens it into CSV/JSON keyed by control_id, with derived
fields: family (AC, AU, etc.), base_control (AC-2 for AC-2(1)), and
enhancement_number when applicable.

Output:
  - 800-53-r5-controls.csv
  - 800-53-r5-controls.json
"""
import csv
import json
import re

import openpyxl


def parse_id(control_id: str):
    """Parse a control identifier into family + base + enhancement.

    Examples:
      AC-1     -> family=AC, base=AC-1, enhancement=None
      AC-2(1)  -> family=AC, base=AC-2, enhancement=1
      PE-3(5)  -> family=PE, base=PE-3, enhancement=5
    """
    m = re.match(r"^([A-Z]{2})-(\d+)(?:\((\d+)\))?", control_id.strip())
    if not m:
        return "", control_id, None
    family = m.group(1)
    number = m.group(2)
    enhancement = int(m.group(3)) if m.group(3) else None
    base = f"{family}-{number}"
    return family, base, enhancement


def clean(s):
    """Normalize cell text: strip, normalize whitespace."""
    if s is None:
        return ""
    return re.sub(r"\s+\n", "\n", str(s)).strip()


def parse_related(s: str) -> list:
    """Parse the 'Related Controls' cell into a list of identifiers."""
    s = clean(s)
    if not s or s.lower() in ("none", "n/a"):
        return []
    # Comma- and semicolon-separated; preserve enhancement parens
    parts = re.split(r"[,;]\s*", s)
    return [p.strip() for p in parts if p.strip()]


def main():
    wb = openpyxl.load_workbook("800-53-r5-controls.xlsx", data_only=True)
    ws = wb[wb.sheetnames[0]]  # "SP 800-53 Revision 5"

    rows = []
    for r in range(2, ws.max_row + 1):
        cid_raw = ws.cell(r, 1).value
        if not cid_raw:
            continue
        cid = clean(cid_raw)
        family, base, enhancement = parse_id(cid)
        rows.append({
            "control_id": cid,
            "family": family,
            "base_control": base,
            "enhancement": enhancement if enhancement is not None else "",
            "is_enhancement": enhancement is not None,
            "name": clean(ws.cell(r, 2).value),
            "control_text": clean(ws.cell(r, 3).value),
            "discussion": clean(ws.cell(r, 4).value),
            "related_controls": parse_related(ws.cell(r, 5).value),
        })

    # Sort: by family, then base number, then enhancement number
    def sort_key(r):
        fam = r["family"]
        m = re.match(r"^[A-Z]{2}-(\d+)", r["base_control"])
        base_num = int(m.group(1)) if m else 999
        enh = r["enhancement"] if r["enhancement"] != "" else 0
        return (fam, base_num, enh)

    rows.sort(key=sort_key)

    # CSV: flatten related_controls as a semicolon-separated string
    csv_rows = []
    for r in rows:
        csv_row = {**r, "related_controls": "; ".join(r["related_controls"])}
        csv_rows.append(csv_row)

    fields = ["control_id", "family", "base_control", "enhancement",
              "is_enhancement", "name", "control_text", "discussion",
              "related_controls"]

    with open("800-53-r5-controls.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(csv_rows)

    with open("800-53-r5-controls.json", "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)

    # Summary
    from collections import Counter
    by_family = Counter(r["family"] for r in rows)
    n_base = sum(1 for r in rows if not r["is_enhancement"])
    n_enh = sum(1 for r in rows if r["is_enhancement"])
    print(f"Total entries: {len(rows)} ({n_base} controls + {n_enh} enhancements)")
    print(f"Families ({len(by_family)}):")
    for fam, n in sorted(by_family.items()):
        print(f"  {fam}: {n}")
    print(f"\nFirst 3 entries:")
    for r in rows[:3]:
        print(f"  {r['control_id']}: {r['name']} (family={r['family']}, related={len(r['related_controls'])})")


if __name__ == "__main__":
    main()
