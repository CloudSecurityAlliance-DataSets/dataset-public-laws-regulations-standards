#!/usr/bin/env python3
"""Extract ENX ISA v6 (TISAX) controls from the official XLSX into CSV/JSON.

Input: isa-6-en.xlsx — the ENX Information Security Assessment v6
catalog. Three control sheets:
  - 'Information Security' — IS controls
  - 'Prototype Protection' — PP additional requirements (sections 8.x)
  - 'Data Protection' — DP additional requirements (sections 9.x)

Each sheet has the same column layout (90+ columns). Control rows are
identifiable by Row_Format='#REF!' in column 1.

Output: one row per control, with the key text fields collected.
"""
import csv
import json

import openpyxl

# Map source column number -> output field name. Aligned with row-2 headers.
FIELDS = {
    3: "control_id",
    8: "control_question",
    9: "objective",
    10: "requirements_must",
    11: "requirements_should",
    12: "additional_requirements_high",
    13: "additional_requirements_very_high",
    14: "additional_requirements_sga",
    15: "usual_responsible_role",
    16: "mappings_other_standards",
    17: "mappings_implementation_guidance",
    18: "measures_recommendations",
    23: "further_information",
    24: "examples_normal_protection",
    25: "examples_high_protection",
    26: "examples_very_high_protection",
    27: "possible_questions",
    28: "possible_evidence",
}


def clean(s):
    """Normalize cell text: strip, collapse non-breaking spaces."""
    if s is None:
        return ""
    s = str(s).replace("\xa0", " ").strip()
    return s


def extract_sheet(ws, sheet_name):
    """Walk a sheet and return (control_rows, section_headers).

    Tracks header rows so each control row carries the section_id /
    section_title context from its parent header(s).
    """
    rows = []
    current_section_id = ""
    current_section_title = ""

    for r in range(3, ws.max_row + 1):
        row_format = clean(ws.cell(r, 1).value)
        is_title = clean(ws.cell(r, 2).value)
        control_id = clean(ws.cell(r, 3).value)

        if row_format == "header" and control_id:
            # Header row: track section context
            # Header title is usually in col 8 (Control question column)
            title = clean(ws.cell(r, 8).value)
            if is_title in ("1", "2"):
                current_section_id = control_id
                current_section_title = title
            continue

        if row_format == "#REF!" and control_id:
            # Control row: extract fields
            entry = {
                "sheet": sheet_name,
                "section_id": current_section_id,
                "section_title": current_section_title,
            }
            for col_num, field_name in FIELDS.items():
                if col_num == 3:
                    entry["control_id"] = control_id
                else:
                    entry[field_name] = clean(ws.cell(r, col_num).value)
            rows.append(entry)

    return rows


def main():
    wb = openpyxl.load_workbook("isa-6-en.xlsx", data_only=True)
    all_rows = []
    for sheet_name in ["Information Security", "Prototype Protection", "Data Protection"]:
        ws = wb[sheet_name]
        rows = extract_sheet(ws, sheet_name)
        print(f"{sheet_name}: {len(rows)} controls")
        all_rows.extend(rows)

    print(f"\nTotal: {len(all_rows)} controls")

    # Field order for CSV (consistent across sheets)
    fields = ["sheet", "control_id", "section_id", "section_title"] + [
        v for v in FIELDS.values() if v != "control_id"
    ]

    with open("isa-6.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(all_rows)

    with open("isa-6.json", "w", encoding="utf-8") as f:
        json.dump(all_rows, f, indent=2, ensure_ascii=False)

    # Distribution check
    from collections import Counter
    by_sheet = Counter(r["sheet"] for r in all_rows)
    print(f"By sheet: {dict(by_sheet)}")
    # Sample
    if all_rows:
        r = all_rows[0]
        print(f"\nFirst control ({r['control_id']}):")
        print(f"  Section: {r['section_id']} - {r['section_title'][:60]}")
        print(f"  Question: {r['control_question'][:80]}")


if __name__ == "__main__":
    main()
