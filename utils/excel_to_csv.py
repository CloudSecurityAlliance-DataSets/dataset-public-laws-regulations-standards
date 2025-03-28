#!/usr/bin/env python3.12
"""
Excel to CSV Converter

This script converts Excel files (.xlsx) to CSV format, properly handling merged cells.
For each input Excel file, it creates a directory with the same name as the Excel file,
and within that directory, it creates one CSV file for each worksheet in the Excel file.

Usage:
    python excel_to_csv.py --input <excel_file_path>

Example:
    python excel_to_csv.py --input FedRAMP-POAM-Template.xlsx
"""

import argparse
import os
import re
import pandas as pd
import openpyxl


def sanitize_filename(name):
    """
    Sanitize a string to be used as a valid filename.
    
    Args:
        name (str): The input string to sanitize
        
    Returns:
        str: A sanitized string with invalid filename characters replaced by underscores
    """
    # Replace invalid filename characters with underscores
    # Windows and many filesystems don't allow these characters: \ / * ? : " < > |
    return re.sub(r'[\\/*?:"<>|]', '_', name)


def excel_to_csv(input_file):
    """
    Convert an Excel file to multiple CSV files (one per sheet).
    
    Args:
        input_file (str): Path to the input Excel file
        
    Returns:
        None: Creates CSV files in a directory named after the input file
    
    Notes:
        - Creates a directory named after the Excel file (without extension)
        - Handles merged cells by replicating their value across all merged coordinates
        - CSV files are named as: <ExcelFileName>-<SheetName>.csv
    """
    # STEP 1: Create output directory from input filename
    base_name = os.path.basename(input_file)
    output_dir = os.path.splitext(base_name)[0]
    sanitized_dir = sanitize_filename(output_dir)
    
    # Create output directory if it doesn't exist
    if not os.path.exists(sanitized_dir):
        os.makedirs(sanitized_dir)
        print(f"Created output directory: {sanitized_dir}")
    
    # STEP 2: Load the Excel file with openpyxl to handle merged cells
    # data_only=True ensures we get values, not formulas
    print(f"Loading Excel file: {input_file}")
    wb = openpyxl.load_workbook(input_file, data_only=True)
    
    # STEP 3: Process each worksheet
    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        # Get the worksheet
        ws = wb[sheet_name]
        
        # Initialize list to store row data
        data = []
        
        # STEP 3.1: Handle merged cells
        # Create a mapping of all cells in merged ranges to the corresponding merged cell value
        merged_cells = ws.merged_cells.ranges
        merged_cell_coords = {}
        
        # Map each cell within a merged range to the value from the top-left cell of the merge
        for merged_range in merged_cells:
            # Get the value from the top-left cell of the merged range
            top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
            # Map every cell in the merged range to this value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cell_coords[(row, col)] = top_left_value
        
        # STEP 3.2: Extract headers from the first row
        headers = []
        for col in range(1, ws.max_column + 1):
            # Check if header cell is part of a merged range
            if (1, col) in merged_cell_coords:
                cell_value = merged_cell_coords[(1, col)]
            else:
                cell_value = ws.cell(1, col).value
            
            # Use the cell value if available, otherwise use a generic column name
            headers.append(cell_value if cell_value is not None else f"Col_{col}")
        
        # STEP 3.3: Process each data row (skip header row)
        for row in range(2, ws.max_row + 1):  # Start from row 2 (after header)
            row_data = {}
            for col in range(1, ws.max_column + 1):
                # Check if this cell is in a merged range
                if (row, col) in merged_cell_coords:
                    cell_value = merged_cell_coords[(row, col)]
                else:
                    cell_value = ws.cell(row, col).value
                
                # Add to row data using header as key
                row_data[headers[col-1]] = cell_value
            
            data.append(row_data)
        
        # STEP 3.4: Create DataFrame from extracted data
        df = pd.DataFrame(data)
        
        # STEP 3.5: Save to CSV file
        # Sanitize sheet name for filename
        sanitized_sheet = sanitize_filename(sheet_name)
        
        # Create output filename: <dirname>/<dirname>-<sheetname>.csv
        output_file = f"{sanitized_dir}/{sanitized_dir}-{sanitized_sheet}.csv"
        
        # Write DataFrame to CSV
        df.to_csv(output_file, index=False)
        
        print(f"  → Converted sheet '{sheet_name}' to {output_file}")


def main():
    """
    Main function: Parse command line arguments and call excel_to_csv.
    """
    parser = argparse.ArgumentParser(
        description='Convert Excel sheets to CSV files, properly handling merged cells.',
        epilog='Creates a directory with the name of the Excel file and outputs one CSV per sheet.'
    )
    parser.add_argument('--input', required=True, help='Path to the input Excel file')
    
    args = parser.parse_args()
    
    # Validate that input file exists
    if not os.path.exists(args.input):
        print(f"Error: Input file '{args.input}' does not exist.")
        return 1
    
    # Convert Excel to CSV
    excel_to_csv(args.input)
    return 0


if __name__ == "__main__":
    main()
